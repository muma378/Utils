# -*- coding: utf-8 -*-
# dbf2xlsx.py - usage: python dbf2xlsx.py root_dir config.json
# reads dbf files under root_dir and the config.json, 
# removes data if the value of the fields which in the 'check' is 0,
# calculates the average values of fields which in the 'average' and generates 
# an excel contain all files
# author: xiaoyang <xiaoyang0117@gmail.com>
# date: 2016.01.10
import os
import sys
import re
import shutil
import openpyxl as px
from dbfread import DBF

CHECK_FIELDS = ['W15000105', 'W15018103']
# CHECK_COLS = ['U', 'AK']
AVERAGE_FIELDS = ['W19020104', 'W19020102', 'W19020103', 'W19020101', 'W15000105', 'W15018103', 'W15011004', 'W15011005', 'W15011006', 'W15011007', 'W13050002']
# AVERAGE_COLS = ['F', 'G', 'H', 'I', 'U', 'AK', 'AQ', 'AR', 'AS', 'AT', 'CP']
FIELDS_TEXT = {'W19020104': u'土压 下', 'W19020102': u'土压 左', 'W19020103': u'土压 右', 'W19020101': u'土压 上', 'W15000105': u'总推力', 'W15018103': u'刀盘转速', 'W15011004': u'注浆 右上', 'W15011005': u'注浆 右下', 'W15011006': u'注浆 左下', 'W15011007': u'注浆 左上', 'W13050002': u'刀盘扭矩'}
# WORDS_INSERT = {'F': u'土压 下', 'G': u'土压 左', 'H': u'土压 右', 'I': u'土压 上', 'U': u'总推力', 'AK': u'刀盘转速', 'AQ': u'注浆 右上', 'AR': u'注浆 右下', 'AS': u'注浆 左下', 'AT': u'注浆 左上', 'CP': u'刀盘扭矩'}

SRC_SUFFIX = '.dbf'
DST_SUFFIX = '.xlsx'

# NAME_PATTERN = '.*/database/process/p(?P<ring_no>\d+)\.dbf'
NAME_PATTERN = '.*\\\\database\\\\process\\\\p(?P<ring_no>\d+)\.dbf'
RING_TITLE = '环号'
JSON_KEYS = ["check", "average", "text", "append"]
OA_KEY = 'ordered_average'
OT_KEY = 'ordered_text'
OAP_KEY = 'ordered_append'

LOGFILE = 'dbf2xlsx.log'
LOGFP = None
CACHEFILE = 'cache.csv'
CACHEFP = None

def setup(root_dir):
	global LOGFP, CACHEFP
	LOGFP = open(os.path.join(root_dir, LOGFILE), 'a')
	cache_path = os.path.join(root_dir, CACHEFILE)
	# if already exists, created a fake one to read
	if os.path.exists(cache_path):
		shutil.copy(cache_path, root_dir+os.sep+'~'+CACHEFILE)
	CACHEFP = open(cache_path, 'a')

def teardown(root_dir):
	LOGFP.close()
	CACHEFP.close()

def loginfo(msg):
	print(msg)
	LOGFP.write(msg+'\n')

# saves the average value to a csv file to avoid calculating from the start
def cache(fullpath, ring_no, record):
	items = [fullpath, ring_no]+record
	line = ''
	for item in items:
		line += str(item).strip() + ','
	CACHEFP.write(line+'\n')


def loads(root_dir):
	print("Loading cache ...")
	cached_dict = {}
	avg_xlsx = {}
	fake_cache = root_dir+os.sep+'~'+CACHEFILE
	# fake cache file
	if os.path.exists(fake_cache):
		with open(fake_cache, 'r') as f:
			for line in f:
				l = line.split(',')
				cached_dict[l[0]] = True
				avg_xlsx.setdefault(os.path.dirname(l[0]), {}).setdefault(l[1], l[2:])
	return cached_dict, avg_xlsx

class InvalidLineException(Exception):
	pass
		

def loadsettings(settings_file):
	print("Loading settings ... ")
	try:
		with open(settings_file) as f:
			settings = eval(f.read())
			# check if all keys are written correctly
			for key in JSON_KEYS:
				settings[key]
			# ordered attributes need to get average value
			settings[OA_KEY] = sorted(settings["average"].values())
			# ordered text according to attribute above
			settings[OT_KEY] = [ settings["text"][k] for k in settings[OA_KEY] ]
			settings[OAP_KEY] = sorted(settings["append"].values())
	except Exception as e:
		print("Error: unable to parse the setting file, please check %s " % settings_file)
		sys.exit(0)
	return settings


# data structure for avg_xlsx: 
# avg_xlsx = { 
# 'root_dir/parent_dir/a' : { 
# 	'0001': [0.5, 0.6, 2.0, ...]
# 	'0002': [...]
# 	}
# 'root_dir/parent_dir/b' : { ...
# }
def readfiles(root_dir, settings_file):
	items = {}
	cached, avg_xlsx = loads(root_dir)
	rn_parser = re.compile(NAME_PATTERN)

	settings = loadsettings(settings_file)
	for dirpath, dirnames, filenames in os.walk(root_dir):
		for filename in filenames:
			# only need to process the data under dir 'process'
			fullpath = os.path.join(dirpath, filename)
			if cached.get(fullpath, False):
				print('%s was processed before, ignored' % fullpath)
			else:
				r = rn_parser.match(fullpath)
				if r:
					try:
						fields_avg = process(fullpath, settings)
					except ValueError as e:
						loginfo('Error: unable to process %s' % fullpath)
						continue
					# create one sheet for all files in the same folder
					ring_no = r.group('ring_no')
					cache(fullpath, ring_no, fields_avg)
					avg_xlsx.setdefault(dirpath, {}).setdefault(ring_no, fields_avg)

	print('Converting finnished')
	gen_avgxlsx(avg_xlsx, root_dir, settings)
	return items



# reads dbf and removes unqualified data
def process(filepath, settings):
	print("Reading %s ..." % filepath)
	t = DBF(filepath, load=True)
	fields_avg = [ 0.0 for i in settings[OA_KEY]]
	xlsx_list = []
	# row index and data
	print("Filtering data now")
	for ri, record in enumerate(t.records):
		try:
			for attr in settings['check'].values():
				if record[attr] == 0:
					t.deleted.append(record)
					raise InvalidLineException('Warning: line %d in %s is invalid, has been removed ' % (ri, filepath))			
			for fi, avg_attr in enumerate(settings[OA_KEY]):
				fields_avg[fi] += record[avg_attr]

			xlsx_list.append(record)
		except InvalidLineException as e:
			pass

	n_del, n_save = len(t.deleted), len(xlsx_list)
	print('Processing finished, %d in total, %d rows are removed, %d rows are saved.' % (n_del+n_save, n_del, n_save))
	gen_xlsx(xlsx_list, filepath, settings)
	
	fields_avg = [ v / n_save for v in fields_avg ]
	# besides, append the last values for several columns
	for attr in settings[OAP_KEY]:
		fields_avg.append(record[attr])
	return fields_avg


def gen_xlsx(xlsx_list, filepath, settings):
	xlsx_path = filepath.replace(SRC_SUFFIX, DST_SUFFIX)
	print('Saving data to %s now ...' % xlsx_path)
	sheet_name = os.path.basename(xlsx_path).replace(DST_SUFFIX, '')
	wb = px.Workbook(write_only=True)
	ws = wb.create_sheet()
	ws.title = sheet_name

	def append_header(sheet, row):
		headers = [['', ], ['', ]]
		for item in row:
			headers[0].append(item)
			headers[1].append(settings['text'].get(item, ''))
		for header in headers:
			sheet.append(header)

	append_header(ws, xlsx_list[0])
	for counter, row in enumerate(xlsx_list, start=1):
		ws.append([counter] + [ val for val in row.values() ])

	wb.save(xlsx_path)


# name as AVG_rootdir.xlsx
def gen_avgxlsx(avg_xlsx, root_dir, settings):
	avgxlsx_name = 'AVG' + DST_SUFFIX
	avgxlsx_path = os.path.join(root_dir, avgxlsx_name)
	print('Saving average values for %s to %s now ...' % (root_dir, avgxlsx_path))
	wb = px.Workbook(write_only=True)
	for path, data in avg_xlsx.items():
		ws = wb.create_sheet()
		try:
			ws.title = path.replace(os.sep, '-').decode('utf-8')[-47:-17]
		except UnicodeDecodeError as e:
			ws.title =  path.replace(os.sep, '-').decode('gb2312')[-47:-17]
		ws.append([RING_TITLE] + settings[OT_KEY] + settings[OAP_KEY])
		ordered_data = [ [k]+v for k,v in sorted(data.items(), key=lambda x: x[0])]
		for d in ordered_data:
			ws.append(d)

	wb.save(avgxlsx_path)

if __name__ == '__main__':
	root_dir, config = sys.argv[1], sys.argv[2]
	setup(root_dir)
	loginfo("===========================Program Started===========================")
	readfiles(root_dir, config)
	teardown(root_dir)