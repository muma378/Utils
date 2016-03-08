# -*- coding: utf-8 -*-
# sync_hkserv_info.py - to synchronize infomation with respect to resource uploaded in the Hongkong Server
# author: XiaoYang<xiaoyang0117@gmail.com> LiuSize
# date: 2016/02/18
import re
import os
import sys
import urllib2
import threading
import Queue
from urllib import quote
from datetime import datetime
import string

import lxml.html
import openpyxl as px

from configure import task_dict as naming
import log
from timeit import timefunc

RETRIVE_BLOCK_TIMEOUT = 120
WORKER_NUM = 16
FIELDS_NUM = 8
VALID_WORLD = u'合格'
VALID_START_INDEX = 4 	# index starts to stand for valid or not  
SIZE_COL_INDEX = 3 # index of column to represent the size of dir in a project
EXCEL_ID_INDEX = 1 # index of the field of id in summary.xltx
BASE_DIR = os.getcwd()
SUMMARY_TEMPLATE = os.path.join(BASE_DIR, 'templates', 'Summary.xltx')
SUMMARY_EXCEL = os.path.join(BASE_DIR, 'Summary.xlsx')
DETAILS_EXCEL = os.path.join(BASE_DIR, 'Details.xlsx')
SUMMARY_KEY = 'summary'
DETAILS_KEY = 'details'
SUMMARY_EDIT_COLS = string.ascii_uppercase[3:14]	# D ~ N


logger = log.LogHandler('sync.log', stdout=False)


def connect(url, timeout=10, count=3, coding='utf-8'):
	while count > 0:
		try:
			logger.info('connecting to %s' % url)
			response = urllib2.urlopen(url, timeout=timeout)
			return response.read().decode(coding)
		except urllib2.HTTPError, e:
			logger.error('falied to connect to %s, may for %s' % (url, e.reason))
		except urllib2.URLError, e:
			logger.error('unable to open url %s for %s' % (url, e.reason))
		except UnicodeDecodeError, e:
			logger.error('unable to decode %s with %s' % (url, coding))
			raise UnicodeError
		count -= 1
		logger.info('%d times left to reconnect...' % count)
	logger.error('unable to conenct to server %s' % url)
	return None
	# sys.exit(0)

def get_projects_list(url):
	response = connect(url, timeout=30)
	if response == None:
		logger.error('please check your network or the server name provided, exiting now...')
		sys.exit(0)
	doc = lxml.html.document_fromstring(response)
	projects_list = map(lambda x: (x.text_content().strip(), url+quote(x.get('href'), safe='/=?')), doc.xpath('//tr/td[1]/a'))
	return projects_list

class ProjectInfoCrawler(threading.Thread):
	"""threads version to extract info for each project"""
	def __init__(self, queue, out_queue, coding='utf-8'):
		super(ProjectInfoCrawler, self).__init__()
		self.queue = queue
		self.out_queue = out_queue
		self.coding = coding
	
	def run(self):
		while True:
			try:
				dirname, url = self.queue.get(timeout=30)
				
				doc = lxml.html.document_fromstring(connect(url, timeout=600))
				items = doc.xpath('//tbody/tr')
				project_info = []
				for item in items:
					dir_info = map(lambda x: x.strip(), item.text_content().split('\r\n'))
					# to filter out the blank field for each row
					dir_info = filter(lambda x: len(x), dir_info)

					try:
						# remove rows whose fields' number is not correct
						if len(dir_info) == FIELDS_NUM:
							project_info.append(dir_info)
						# data field would be empty if no data in the directory
						elif len(dir_info) == FIELDS_NUM-1 and int(dir_info[SIZE_COL_INDEX])==0:
							# TODO: not sure to append None
							dir_info.append(None)
							project_info.append(dir_info)
						else:
							raise ValueError("incorrect number of fields")
					except (ValueError, IndexError), e:
						content = re.sub('\s{2,}', '\t', item.text_content()).encode(self.coding)
						logger.error('illegal row when parsing %s: %s' % (dirname, content))
				print url
				self.out_queue.put((dirname, project_info))
			except ValueError, e:	# response == None
				logger.error('unable to parse directory %s, ignored' % dirname)
			except UnicodeError, e:	# unable to decode with 'utf-8'
				logger.error(e)
			except Queue.Empty:	# processed finished or blocked in fetching urls
				logger.info('%s exits for empty queue' % self.name)
				break
			self.queue.task_done()


def retrive(queue, data):
	counter = 0
	try:
		while True:
			dirname, info = queue.get(timeout=RETRIVE_BLOCK_TIMEOUT)
			# logger.info(dirname)
			# logger.info(str(info))
			# extract info for a project
			data[SUMMARY_KEY][dirname] = stat(info)
			# get project name instead of id
			data[DETAILS_KEY][dirname] = info
			queue.task_done()
			counter += 1		
	except Queue.Empty:
		logger.info('%d projects are retrived' % counter)

	return data

# sample of info:
# [['G0151', '46.69 M', u'正常', '133', '--', u'合格', '--', '01/05 19:16:26'], 
# ['G0153', '0.00 Byte', u'正常', '0, '--', '--', '--', None]]
def stat(info):  
	summary = {
		'date':datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
		'count':len(info),
		'over450':[0, 0, 0],	# stands for the quality authorized by 3 roles
		'bet400and450':[0, 0, 0],
		'under400':[0, 0, 0]
		}

	region_map = lambda x: 'over450' if x>=450 else 'under400' if x<400 else 'bet400and450'
	
	for item in info:
		valid = summary[region_map(int(item[SIZE_COL_INDEX]))]
		for i in range(len(valid)):
			if item[VALID_START_INDEX+i] == VALID_WORLD:
				valid[i] += 1
	# logger.info(str(summary))
	
	# flatten
	d = summary
	summary_list = [d['date'], d['count']]+d['over450']+d['bet400and450']+d['under400']
	return summary_list

def crawl_info(projects_list, data):
	queue = Queue.Queue()
	out_queue = Queue.Queue()

	for i in range(WORKER_NUM):
		t = ProjectInfoCrawler(queue, out_queue)
		t.setDaemon(True)
		t.start()

	# populate queue with data extracted before
	for project in projects_list:
		queue.put(project)

	data = retrive(out_queue, data)
	# wait on the queue until all has been processed
	queue.join()
	out_queue.join()
	logger.info('infomation for projects are crawled successfully')
	return data


def summarize(summary):
	logger.info('writing to %s now...' % SUMMARY_TEMPLATE)
	wb = px.load_workbook(filename=SUMMARY_TEMPLATE)
	ws = wb.active
	for row in ws.rows:
		project_id = row[EXCEL_ID_INDEX].value
		row_index = row[EXCEL_ID_INDEX].row
		if type(project_id) is long or type(project_id) is int:	# if the cell stands for a integer for id
			row_index = str(row_index)
			try:
				for i,c in enumerate(SUMMARY_EDIT_COLS):
					ws[c+row_index] = summary[str(project_id)][i]
			except KeyError, e:
				logger.error('unable to find project id %d' % project_id)
		elif row_index > 2:
			logger.error('field of id in row %d is not a long or an integer' % row_index)
			
	wb.save(SUMMARY_EXCEL)
	logger.info('summary has been written to %s successfully' % SUMMARY_EXCEL)
			
# [['G0151', '46.69 M', u'正常', '133', '--', u'合格', '--', '01/05 19:16:26'], 
def itemize(detail):
	logger.info('writing to %s now...' % DETAILS_EXCEL)
	title = [u'目录名称', u'大小', u'状态', u'文件量', u'外包商质检', u'质检', u'验收', u'最后上传时间']
	wb = px.Workbook(write_only=True)

	sheets_order = sorted(detail.keys())
	for project_id in sheets_order:
		try:
			sheet_title = naming.get(int(project_id), project_id)
		except ValueError, e:
			logger.error('unable to get the project key for %s' % project_id)
			continue

		ws = wb.create_sheet()
		ws.title = sheet_title
		ws.append(title)
		for row in detail[project_id]:
			try:
				ws.append(row)
			except TypeError, e:
				logger.error('unable to write %s in excel file' % str(row))

	wb.save(DETAILS_EXCEL)
	logger.info('detail has been written to %s successfully' % DETAILS_EXCEL)

@timefunc
def sync(url):
	# to get the list of projects
	print('connecting to the server %s...' % url)
	projects_list = get_projects_list(url)

	# to extract infomation about each project and statistics
	print('extracting information...')
	data = {SUMMARY_KEY: {}, DETAILS_KEY: {}}
	crawl_info(projects_list, data)

	print('writing to summary.xlsx and details.xlsx...')
	# to write the info into excel files respectively
	summarize(data[SUMMARY_KEY])
	itemize(data[DETAILS_KEY])


if __name__ == '__main__':
	sync('http://v-prj-002.cloudapp.net/file/')
	print 'Succeed'