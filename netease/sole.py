#!/usr/bin/env python
# -*- coding: utf-8 -*-
import os
import sys
import locale
import codecs
import argparse
import hashlib
from math import sqrt
import openpyxl as px
import jieba.analyse


def get_system_encoding():
    """
    The encoding of the default system locale but falls back to the given
    fallback encoding if the encoding is unsupported by python or could
    not be determined.  See tickets #10335 and #5846
    """
    try:
        encoding = locale.getdefaultlocale()[1] or 'ascii'
        codecs.lookup(encoding)
    except Exception:
        encoding = 'ascii'
    return encoding


DEFAULT_LOCALE_ENCODING = get_system_encoding()

# generator version of traversing
def traverser(src_dir, dst_dir, target='.wav', err=sys.stderr):
    for dirpath, dirnames, filenames in os.walk(src_dir):
        for filename in filenames:
            if filename.endswith(target):
                try:
                    src_file = os.path.join(dirpath, filename)
                    src_dir_len = len(src_dir) if src_dir.endswith(os.sep) else len(src_dir)+1
                    dst_file = os.path.join(dst_dir, src_file[src_dir_len:])    # should not use replace
                    yield src_file, dst_file
                except Exception as e:
                    err.write("unable to process {filename} for {reason}\n".format(filename=src_file, reason=e))
    
# only return content if its sheet_title, column and row were valid
def load_xlsx(filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.load_workbook(filename=filename)
	workbook = {}
	for ws in wb:
		if sheet_titles and not ws.title in sheet_titles:
			continue
		workbook[ws.title] = []
		for row in ws.rows:
			vals = []
			for cell in row:
				vals.append(cell.value)
			workbook[ws.title].append(vals)
	return workbook


def dump_xlsx(workbook, filename, sheet_titles=[], columns=[], rows=[]):
	wb = px.Workbook()
	actived = False
	if isinstance(workbook, dict):
		for sheet_title, sheet_content in workbook.items():
			if sheet_titles and not sheet_title in sheet_titles:
				continue

			if actived:
				ws = create_sheet(title=sheet_title)
			else:
				ws = wb.active		# have to call active in the first time to create sheet
				ws.title = sheet_title
				actived = True

			for row in sheet_content:
				ws.append(row)
	else:	# syntax sugar for one sheet
		ws = wb.active
		for row in workbook:
			ws.append(row)

	wb.save(filename=filename)


class SimhashObject(object):
	"""a document was taken and a hash object was returned"""
	def __init__(self, doc, label=None):
		super(SimhashObject, self).__init__()
		self.doc = doc
		self.label = label
		self._tokenize()
		self._hash()
		
	def _tokenize(self, topK=20):
		self.tokens_with_weight = jieba.analyse.extract_tags(self.doc, topK=topK, withWeight=True)
		return self.tokens_with_weight

	# width deciedes how many bits was kept 
	# and need to be longer than hash_fn's block size
	def _hash(self, hash_fn=hashlib.md5, width=128):

		def token_sign(token, weight):
			hexstr = hash_fn(token).hexdigest()
			binstr = bin(int(hexstr, 16))[2:].zfill(width)	# get the binary representation
			sign = map(lambda x: (int(x)-0.5)*weight, list(binstr))	# flip the weight if the bit was zero
			return sign

		def list_add(list1, list2):
			"""
			add corresponding elements in each list
			"""
			assert len(list1) == len(list2)
			return map(lambda x: x[0]+x[1], zip(list1, list2))


		signs = map(lambda x: token_sign(x[0].encode('utf-8'), x[1]), self.tokens_with_weight)	# get hash sign for each token
		fingerprint = reduce(lambda x, y: list_add(x, y), signs)	# add each bit in all signs
		self.fp_bits = map(lambda x: 1 if x > 0 else 0, fingerprint)
		self.fingerprint = ''.join(map(lambda x: str(x), self.fp_bits))
		return self.fingerprint

	# for its speciality (only including 0 and 1), 
	# its cosine value (Va*Vb/(|Va|*|Vb|)) can be computed in a trick way
	def cosine_sim(self, that):
		fp_dot = map(lambda x: x[0] * x[1], zip(self.fp_bits, that.fp_bits))	# get its bit_and
		
		count = lambda l: len(filter(lambda x: x, l)) * 1.0

		numerator = count(fp_dot)
		denominator = count(self.fp_bits)*count(that.fp_bits)

		assert denominator != 0
		return sqrt(numerator/denominator)


	def hamming_dist(self, that):
		fp_dot = map(lambda x: 1 if x[0]!=x[1] else 0, zip(self.fp_bits, that.fp_bits))	# xor
		return len(filter(lambda x: x, fp_dot))
	

def check_content_similarity(titles, content, args, errors):
	try:
		content_col = titles.index(args.column)
	except ValueError as e:
		print("未能在'%s'中找到名称为%s的列" % (args.filepath, args.column.encode(DEFAULT_LOCALE_ENCODING)))
		sys.exit(1)

	simhash_objects = []
	print "generating simhash objects - %d in total" % len(content)
	for i, row in enumerate(content):
		if len(row) > content_col:
			if row[content_col] and isinstance(row[content_col], unicode):
				so = SimhashObject(row[content_col], label=(i+1))
				simhash_objects.append(so)

	for i, so1 in enumerate(simhash_objects):
		if i == len(simhash_objects):
			break
		for so2 in simhash_objects[i+1:]:
			distance = so1.hamming_dist(so2)
			if distance < args.distance:
				print "{0} 和 {1} 行相似，汉明距为 {2}".format(so1.label, so2.label, distance)
				errors[so1.label-1] += u"与%d行相似，simhash的汉明距离为%f" % (so2.label+1, distance)
				break	# no more need to compare the similarity between so1 and the rest

			# similarity = so1.cosine_sim(so2)
			# if similarity > conditions['similarity']:
			# 	print "{0} & {1}: {2}".format(so1.label, so2.label, similarity)
			# 	errors[so1.label-1] += u"与%d行过于相似，simhash的余弦距离为%f" % (so2.label+1, similarity)
			# 	break	# no more need to compare the similarity between so1 and the rest

	return errors


def export_error_message(titles, content, errors):
	verbose = [titles + [u"错误信息"]]
	for row, error in zip(content, errors[:-1]):
		verbose.append(row + [error])
	global_error = ['' for x in titles] + [errors[-1]]
	verbose.append(global_error)
	return verbose


def process_sheet(sheet, args, start=0):
	titles, content = sheet[start], sheet[start+1:100]	# TEST

	# initialize, one for each row, extra comment for global error
	errors = ['' for i in content] + ['']

	# checks the similarity for each pairs
	errors = check_content_similarity(titles, content, args, errors)

	# exports to 2 files, one only for the valid and one for all and reason 
	return export_error_message(titles, content, errors)


def process(filepath, args):
	print("processing {0} ...".format(os.path.basename(filepath)))
	workbook = load_xlsx(filepath)
	args.filepath = filepath

	verbose_workbook = {}
	for title, sheet in workbook.items():
		verbose_workbook[title] = process_sheet(sheet, args)

	verbose_filename = "%s-verbose%s" % os.path.splitext(filepath)

	if os.path.exists(verbose_filename):
		os.remove(verbose_filename)

	dump_xlsx(verbose_workbook, verbose_filename)
	print("exported %s" % verbose_filename)


def main():
	parser = argparse.ArgumentParser(description='Finds near-duplicates text.')
	parser.add_argument('-p', '--path', help='Path to the target file(s).')
	parser.add_argument('-d', '--distance', type=int, default=20, metavar='value',
		help='Hamming distance representing the difference between 2 text.')
	parser.add_argument('-c', '--column', default=u'新闻内容', metavar='name',
		help='The tag name indicates the column to compare.'
		)
	args = parser.parse_args(sys.argv[1:])

	if not isinstance(args.column, unicode):
		args.column = args.column.decode(DEFAULT_LOCALE_ENCODING)

	if os.path.exists(args.path):
		if os.path.isfile(args.path):
			process(args.path, args)
		else:
			for src_file, _ in traverser(args.path, '', target='.xlsx'):
				if os.path.basename(src_file).startswith('~'):	# temp files for xlsx
					continue
				process(src_file, args)
	else:
		print("Specified path is not illeagal")
		

if __name__ == '__main__':
	main()